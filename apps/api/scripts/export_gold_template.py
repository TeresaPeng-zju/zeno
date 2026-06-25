"""Export a gold-labeling TEMPLATE for the extraction labeling function (LF).

Why this exists
---------------
The whole skill-evidence ledger (app/data/jd_evidence.json) is produced by ONE
labeling function: the keyword table `SKILL_KEYWORDS` in
`scripts/build_jd_evidence.py`. Today that LF is **unmeasured** — we trust it on
faith. raw/README.md §4 says it plainly: "没有 gold，加源是凭感觉".

This script samples real JD docs and writes a labeling template so a human can
record the *true* graph skills per doc. Crucially it does **NOT** pre-fill the
LF's own prediction into the `skills` field — pre-filling would leak the system's
answer into the label and re-introduce exactly the confirmation bias we are trying
to measure away. The labeler reads the JD text cold and picks from the 23 graph
skills (see app/data/gold/README.md for the menu).

Output: app/data/gold/extraction_gold.template.jsonl
Then: copy/rename to extraction_gold.jsonl, fill each `skills`, and run
      `python -m app.eval.extraction.extraction_eval`.

Sampling is deterministic (sorted by doc_id = sha1(text)[:8]), so re-running with
the same N yields the same template — reproducible like everything else in Zeno.

Run:
    cd apps/api && python -m scripts.export_gold_template            # default N=25
    cd apps/api && python -m scripts.export_gold_template --n 30
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from pathlib import Path

from scripts.build_jd_evidence import _XLSX, _norm

_API_ROOT = Path(__file__).resolve().parent.parent
_GOLD_DIR = _API_ROOT / "app" / "data" / "gold"
_OUT = _GOLD_DIR / "extraction_gold.template.jsonl"

_SOURCE_ID = os.environ.get("JD_SOURCE_ID", "jd/multi_source")
_JD_COLS = ("职位名称", "职位描述", "职位要求")


def _doc_id(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:8]


def _read_docs() -> list[dict]:
    """Read JD rows into readable per-doc records (original case for humans).

    Uses openpyxl directly (pandas' xlsx engine) to avoid a heavy pandas dep just
    to read three columns.
    """
    from openpyxl import load_workbook  # local import: only this corpus pass needs it

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col_idx = {name: header.index(name) for name in _JD_COLS if name in header}

    def cell(row: tuple, name: str) -> str:
        i = col_idx.get(name)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i])

    docs: list[dict] = []
    for row in rows:
        # Human-readable text (original case, newline-joined). The LF normalizes
        # (lowercase + collapse whitespace) at match time, so case here is moot.
        readable = "\n".join(cell(row, col).strip() for col in _JD_COLS).strip()
        if not readable:
            continue
        # doc_id is computed on the NORMALIZED blob so it is stable regardless of
        # incidental whitespace/case differences in the source.
        norm_blob = " ".join(_norm(cell(row, col)) for col in _JD_COLS)
        docs.append({"doc_id": _doc_id(norm_blob), "text": readable})
    wb.close()
    return docs


def _sample(docs: list[dict], n: int) -> list[dict]:
    """Deterministic, content-addressed sample: sort by doc_id, take first n.

    Sorting by sha1 is effectively a stable pseudo-random shuffle, so we avoid
    cherry-picking easy-to-label docs (raw/README.md §4) while staying reproducible.
    Dedupes identical docs by doc_id.
    """
    seen: dict[str, dict] = {}
    for d in docs:
        seen.setdefault(d["doc_id"], d)
    unique = sorted(seen.values(), key=lambda d: d["doc_id"])
    return unique[: min(n, len(unique))]


def export(n: int) -> Path:
    docs = _sample(_read_docs(), n)
    _GOLD_DIR.mkdir(parents=True, exist_ok=True)
    with _OUT.open("w", encoding="utf-8") as f:
        for d in docs:
            record = {
                "doc_id": d["doc_id"],
                "source_id": _SOURCE_ID,
                "lang": "zh",
                # TO LABEL: pick the graph skills this JD genuinely requires.
                # Left empty on purpose (no LF prediction leaked in). Menu of valid
                # ids: app/data/gold/README.md (only the 23 graph skills are valid).
                "skills": [],
                "note": "",
                "text": d["text"],
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return _OUT


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--n", type=int, default=25, help="docs to sample (15–30 recommended)")
    args = parser.parse_args()

    out = export(args.n)
    n_written = sum(1 for _ in out.open(encoding="utf-8"))
    print(f"Wrote {out.relative_to(_API_ROOT)}  ({n_written} docs to label)")
    print(
        "Next:\n"
        "  1) cp app/data/gold/extraction_gold.template.jsonl "
        "app/data/gold/extraction_gold.jsonl\n"
        "  2) fill each line's \"skills\" using the menu in app/data/gold/README.md\n"
        "  3) python -m app.eval.extraction.extraction_eval"
    )


if __name__ == "__main__":
    main()
